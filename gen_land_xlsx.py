# -*- coding: utf-8 -*-
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

CSV_PATH = r"C:\Users\huaxi\WorkBuddy\Claw\land_monitor_results.csv"
XLSX_PATH = r"C:\Users\huaxi\WorkBuddy\Claw\land_monitor_latest.xlsx"

KEYWORDS = ("住宅", "居住", "普通商品房", "商品房")
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

wb = Workbook()
ws = wb.active
ws.title = "土地公告"

with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
    reader = csv.reader(f)
    rows = list(reader)

header = rows[0]
ws.append(header)
# 表头加粗
for c in ws[1]:
    c.font = Font(bold=True)
    c.alignment = Alignment(vertical="center", wrap_text=True)

yellow_count = 0
for row in rows[1:]:
    if not row or not any(cell.strip() for cell in row):
        continue
    ws.append(row)
    # 标黄判据：行的任一单元格含关键词
    row_has_kw = any(any(kw in (cell or "") for kw in KEYWORDS) for cell in row)
    if row_has_kw:
        yellow_count += 1
        for c in ws[ws.max_row]:
            c.fill = YELLOW
            c.alignment = Alignment(vertical="center", wrap_text=True)

# 列宽自适应（简单设）
widths = [40, 12, 10, 14, 40, 24, 14, 14, 10, 30, 14, 14, 12, 10, 10, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

wb.save(XLSX_PATH)
print(f"XLSX 生成完成: {XLSX_PATH}")
print(f"数据行数: {ws.max_row - 1}")
print(f"标黄行数: {yellow_count}")
